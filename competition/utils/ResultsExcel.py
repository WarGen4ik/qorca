import openpyxl
from openpyxl.styles import *
from django.conf import settings
from django.utils.translation import gettext as _

from auth_main.models import User, Profile
from core.models import CompetitionUser, CompetitionTeam, TeamRelationToUser, Distance, UserDistance


class ResultsExcel:
    alf = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    def __init__(self, competition):
        self.competition = competition

    def create_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        alignment = Alignment(horizontal='center')
        border = Border(
            left=Side(border_style="thin", color='000000'),
            right=Side(border_style="thin", color='000000'),
            top=Side(border_style="thin", color='000000'),
            bottom=Side(border_style="thin", color='000000'),
            outline=Side(border_style="thin", color='000000')
        )
        fill = PatternFill(start_color='eff0f1',
                           end_color='eff0f1',
                           fill_type='solid')
        for day in range(self.competition.count_days):
            if day != 0:
                ws = wb.create_sheet()
            ws.title = _('Day ') + str(day + 1)
            alf_index = 1
            index = 1
            columns = [_('Member'), _('Age group'), _('Team'), _('City'), _('Time'), _('Points')]
            groups = ['25-29', '30-34', '35-39', '40-44', '45-49', '50-54', '55-59', '60-64', '65-69', '70-74', '75-79',
                      '80-84', '85-89', '90-94']
            distance_index = 1
            for distance in Distance.objects.filter(competition=self.competition, day=day + 1).all():
                char = self.get_char(alf_index - 1)
                next_char = self.get_char(alf_index + 4)
                for x in range(4):
                    ws.column_dimensions[self.get_char(x)].width = 15

                for gender in range(2, 0, -1):
                    printed_distance = False
                    if gender == 1:
                        gender_text = _('Males')
                    else:
                        gender_text = _('Females')

                    for group in groups:
                        users_distances = UserDistance.objects \
                            .filter(distance=distance, user__profile__gender=gender, user__profile__age_group=group,
                                    is_finished=True) \
                            .order_by('-points')
                        if not users_distances:
                            continue
                        column_index = 1
                        if not printed_distance:
                            ws.merge_cells('{}{}:{}{}'.format(char, index, next_char, index))
                            ws['{}{}'.format(char, index)].font = Font(size=12, bold=True)
                            ws['{}{}'.format(char, index)].alignment = alignment
                            ws['{}{}'.format(char, index)] = _('Distance ') + '№{}'.format(
                                distance_index) + ' - ' + str(
                                distance.length) + ' ' + distance.get_type_display() + ' ' + gender_text
                            index += 2
                            ws['{}{}'.format(self.get_char(column_index - 1), index)] = columns[column_index - 1]
                            ws['{}{}'.format(self.get_char(column_index - 1), index)].alignment = alignment
                            ws['{}{}'.format(self.get_char(column_index - 1), index)].border = border
                            ws['{}{}'.format(self.get_char(column_index - 1), index)].fill = fill
                            ws['{}{}'.format(self.get_char(column_index), index)] = columns[column_index]
                            ws['{}{}'.format(self.get_char(column_index), index)].alignment = alignment
                            ws['{}{}'.format(self.get_char(column_index), index)].border = border
                            ws['{}{}'.format(self.get_char(column_index), index)].fill = fill
                            ws['{}{}'.format(self.get_char(column_index + 1), index)] = columns[column_index + 1]
                            ws['{}{}'.format(self.get_char(column_index + 1), index)].alignment = alignment
                            ws['{}{}'.format(self.get_char(column_index + 1), index)].border = border
                            ws['{}{}'.format(self.get_char(column_index + 1), index)].fill = fill
                            ws['{}{}'.format(self.get_char(column_index + 2), index)] = columns[column_index + 2]
                            ws['{}{}'.format(self.get_char(column_index + 2), index)].alignment = alignment
                            ws['{}{}'.format(self.get_char(column_index + 2), index)].border = border
                            ws['{}{}'.format(self.get_char(column_index + 2), index)].fill = fill
                            ws['{}{}'.format(self.get_char(column_index + 3), index)] = columns[column_index + 3]
                            ws['{}{}'.format(self.get_char(column_index + 3), index)].alignment = alignment
                            ws['{}{}'.format(self.get_char(column_index + 3), index)].border = border
                            ws['{}{}'.format(self.get_char(column_index + 3), index)].fill = fill
                            ws['{}{}'.format(self.get_char(column_index + 4), index)] = columns[column_index + 4]
                            ws['{}{}'.format(self.get_char(column_index + 4), index)].alignment = alignment
                            ws['{}{}'.format(self.get_char(column_index + 4), index)].border = border
                            ws['{}{}'.format(self.get_char(column_index + 4), index)].fill = fill
                            index += 1
                            printed_distance = True

                        ws.merge_cells('{}{}:{}{}'.format(char, index, next_char, index))
                        ws['{}{}'.format(char, index)].font = Font(size=12, bold=True)
                        ws['{}{}'.format(char, index)].alignment = alignment
                        ws['{}{}'.format(char, index)] = _('Age group') + ' ' + group
                        index += 1

                        ws.merge_cells('{}{}:{}{}'.format(char, index, next_char, index))
                        ws['{}{}'.format(char, index)].font = Font(size=12, bold=True)

                        ws['{}{}'.format(char, index)].alignment = alignment
                        # ws['{}{}'.format(char, index)].border = border
                        index += 1

                        for user_distance in users_distances:
                            ws['{}{}'.format(self.get_char(column_index - 1), index)] = user_distance.user.full_name
                            ws['{}{}'.format(self.get_char(column_index - 1), index)].border = border
                            ws['{}{}'.format(self.get_char(column_index), index)] = '{}({})'.format(
                                user_distance.user.profile.get_age_group(),
                                user_distance.user.profile.get_age_group_numbers())
                            ws['{}{}'.format(self.get_char(column_index), index)].border = border
                            try:
                                team = TeamRelationToUser.objects.filter(user=user_distance.user).first().team
                                CompetitionTeam.objects.get(team=team, competition=self.competition, is_complete=True)
                                team = team.name
                            except:
                                team = user_distance.user.profile.get_default_team()
                            ws['{}{}'.format(self.get_char(column_index + 1), index)] = team
                            ws['{}{}'.format(self.get_char(column_index + 1), index)].border = border
                            ws['{}{}'.format(self.get_char(column_index + 2), index)] = user_distance.user.profile.city
                            ws['{}{}'.format(self.get_char(column_index + 2), index)].border = border
                            ws['{}{}'.format(self.get_char(column_index + 3), index)] = user_distance.result_time
                            ws['{}{}'.format(self.get_char(column_index + 3), index)].border = border
                            ws['{}{}'.format(self.get_char(column_index + 4), index)] = user_distance.points
                            ws['{}{}'.format(self.get_char(column_index + 4), index)].border = border
                            index += 1
                        index += 1
                distance_index += 1
        path = settings.BASE_DIR + "/media/results/" + str(self.competition.id) + ".xlsx"
        wb.save(path)
        return path

    def create_rating(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        alignment = Alignment(horizontal='center')
        border = Border(
            left=Side(border_style="thin", color='000000'),
            right=Side(border_style="thin", color='000000'),
            top=Side(border_style="thin", color='000000'),
            bottom=Side(border_style="thin", color='000000'),
            outline=Side(border_style="thin", color='000000')
        )
        border_medium = Border(
            left=Side(border_style="medium", color='000000'),
            right=Side(border_style="medium", color='000000'),
            top=Side(border_style="medium", color='000000'),
            bottom=Side(border_style="medium", color='000000'),
            outline=Side(border_style="medium", color='000000')
        )
        fill = PatternFill(start_color='eff0f1',
                           end_color='eff0f1',
                           fill_type='solid')
        columns = ['№', _('Member'), _('Age group'), _('Team'), _('City'), _('Points')]
        index = 1
        char = self.get_char(0)
        next_char = self.get_char(5)

        for x in range(1, 5, 1):
            ws.column_dimensions[self.get_char(x)].width = 15

        for gender in range(2, 0, -1):
            if gender == 1:
                gender_text = _('males')
            else:
                gender_text = _('females')

            users_ids = UserDistance.objects.filter(
                distance__competition=self.competition,
                is_finished=True,
                user__profile__gender=gender) \
                .values('user').distinct()

            if not users_ids:
                continue

            ws.merge_cells('{}{}:{}{}'.format(char, index, next_char, index))
            ws['{}{}'.format(char, index)].font = Font(size=12, bold=True)
            ws['{}{}'.format(char, index)].alignment = alignment
            ws['{}{}'.format(char, index)].border = border
            ws['{}{}'.format(char, index)] = _('Rating in absolute between %(gender)s') % {'gender': gender_text}
            index += 1
            column_index = 0
            for text in columns:
                ws['{}{}'.format(self.get_char(column_index), index)] = text
                ws['{}{}'.format(self.get_char(column_index), index)].alignment = alignment
                ws['{}{}'.format(self.get_char(column_index), index)].border = border_medium
                ws['{}{}'.format(self.get_char(column_index), index)].fill = fill
                column_index += 1

            index += 1

            users_points = []
            for user_id in users_ids:
                user_id = user_id['user']
                user_distances = UserDistance.objects.filter(distance__competition=self.competition, user__id=user_id, is_finished=True) \
                                     .exclude(points__isnull=True) \
                                     .order_by('-points')[:3]

                if user_id == 36:
                    user_id = user_id

                points_sum = 0
                if user_distances:
                    for user_distance in user_distances:
                        points_sum += 0 if user_distance.points is None else user_distance.points
                    try:
                        users_points.append({'user': user_distance.user, 'points': points_sum})
                    except:
                        users_points.append({'user': User.objects.filter(id=user_id).first(), 'points': 0})

            users_points = sorted(users_points, key=lambda k: k['points'], reverse=True)
            user_index = 1
            temp = []
            for user_points in users_points:
                user = user_points['user']
                if user.full_name in temp:
                    pass
                temp.append(user.full_name)
                points = user_points['points']

                ws['{}{}'.format(self.get_char(0), index)] = user_index
                ws['{}{}'.format(self.get_char(0), index)].border = border

                ws['{}{}'.format(self.get_char(1), index)] = user.full_name
                ws['{}{}'.format(self.get_char(1), index)].border = border

                ws['{}{}'.format(self.get_char(2), index)] = '{}({})'.format(
                    user.profile.get_age_group(),
                    user.profile.get_age_group_numbers())
                ws['{}{}'.format(self.get_char(2), index)].border = border
                ws['{}{}'.format(self.get_char(2), index)].alignment = alignment
                try:
                    team = TeamRelationToUser.objects.filter(user=user).first().team
                    CompetitionTeam.objects.get(team=team, competition=self.competition, is_complete=True)
                    team = team.name
                except:
                    team = user_distance.user.profile.get_default_team()
                ws['{}{}'.format(self.get_char(3), index)] = team
                ws['{}{}'.format(self.get_char(3), index)].border = border
                ws['{}{}'.format(self.get_char(3), index)].alignment = alignment
                ws['{}{}'.format(self.get_char(4), index)] = user.profile.city
                ws['{}{}'.format(self.get_char(4), index)].border = border
                ws['{}{}'.format(self.get_char(4), index)].alignment = alignment
                ws['{}{}'.format(self.get_char(5), index)] = points
                ws['{}{}'.format(self.get_char(5), index)].border = border
                ws['{}{}'.format(self.get_char(5), index)].alignment = alignment
                index += 1
                user_index += 1
            index += 3

        path = settings.BASE_DIR + "/media/results/" + str(self.competition.id) + "_rating.xlsx"
        wb.save(path)
        return path

    def is_group_exists(self, group, members):
        for member in members:
            if member.profile.get_age_group() == group:
                return True

        return False

    def get_char(self, index):
        try:
            return self.alf[index]
        except:
            index -= 26
            return 'A' + self.alf[index]

    def get_all_members(self):
        competition_rel_user = CompetitionUser.objects.filter(competition=self.competition).all()
        single_members = list()
        for rel in competition_rel_user:
            single_members.append(rel.user)

        teams = CompetitionTeam.objects.filter(competition=self.competition).all()
        team_members = list()

        for team in teams:
            rels = TeamRelationToUser.objects.filter(team=team.team).all()
            for rel in rels:
                team_members.append(rel.user)

        return single_members + list(set(team_members) - set(single_members))
