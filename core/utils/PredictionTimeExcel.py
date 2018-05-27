import openpyxl
from openpyxl.styles import *
from django.conf import settings
from django.utils.translation import gettext as _

from auth_main.models import User, Profile
from competition.utils import time_to_str
from core.models import CompetitionUser, CompetitionTeam, TeamRelationToUser, Distance, UserDistance


class PredictionTimeExcel:
    alf = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    def __init__(self, competition):
        self.competition = competition

    def create_excel(self, is_finished):
        wb = openpyxl.Workbook()
        ws = wb.active
        alignment = Alignment(horizontal='center')
        border = Border(
            left=Side(border_style="thin", color='000000'),
            right=Side(border_style="thin", color='000000'),
            top=Side(border_style="thin", color='000000'),
            bottom=Side(border_style="thin", color='000000'),
            outline=Side(border_style="thin", color='000000')
        )
        for day in range(self.competition.count_days):
            if day != 0:
                ws = wb.create_sheet()
            ws.title = _('Day ') + str(day+1)
            alf_index = 1
            index = 1
            columns = [_('Member'), _('Age group'), _('Team'), _('City'), _('Time'), _('Track')]
            distance_index = 1
            swim_index = 1
            for distance in Distance.objects.filter(competition=self.competition, day=day+1).all():
                char = self.get_char(alf_index - 1)
                next_char = self.get_char(alf_index + 4)
                for x in range(4):
                    ws.column_dimensions[self.get_char(x)].width = 15

                for gender in range(2, 0, -1):
                    if gender == 1:
                        gender_text = _('Males')
                    else:
                        gender_text = _('Females')
                    not_end = True
                    column_index = 1
                    distance_swim_index = 1
                    while not_end:
                        if is_finished:
                            users_distances = UserDistance.objects.filter(distance=distance, user__profile__gender=gender, is_finished=is_finished)\
                            .order_by('-pre_time')[(distance_swim_index - 1) * self.competition.track_count:distance_swim_index * self.competition.track_count]
                        else:
                            users_distances = UserDistance.objects.filter(distance=distance,
                                                                          user__profile__gender=gender) \
                                                  .order_by('-pre_time')[(distance_swim_index - 1) * self.competition.track_count:distance_swim_index * self.competition.track_count]
                        if not users_distances:
                            not_end = False
                            break
                        ws.merge_cells('{}{}:{}{}'.format(char, index, next_char, index))
                        ws['{}{}'.format(char, index)].font = Font(size=14, bold=True)
                        ws['{}{}'.format(char, index)].alignment = alignment
                        ws['{}{}'.format(char, index)] = _('Swim ') + '№{} - {}'.format(swim_index, gender_text)
                        index += 1
                        ws.merge_cells('{}{}:{}{}'.format(char, index, next_char, index))
                        ws['{}{}'.format(char, index)].font = Font(size=12, bold=True)
                        ws['{}{}'.format(char, index)].alignment = alignment
                        ws['{}{}'.format(char, index)] = _('Distance ') + '№{}'.format(distance_index) + ' - ' + str(
                            distance.length) + ' ' + distance.get_type_display()
                        index += 2

                        ws.merge_cells('{}{}:{}{}'.format(char, index, next_char, index))
                        ws['{}{}'.format(char, index)].font = Font(size=12, bold=True)

                        ws['{}{}'.format(char, index)].alignment = alignment
                        # ws['{}{}'.format(char, index)].border = border
                        index += 1
                        ws['{}{}'.format(self.get_char(column_index - 1), index)] = columns[column_index-1]
                        ws['{}{}'.format(self.get_char(column_index - 1), index)].alignment = alignment
                        ws['{}{}'.format(self.get_char(column_index - 1), index)].border = border
                        ws['{}{}'.format(self.get_char(column_index), index)] = columns[column_index]
                        ws['{}{}'.format(self.get_char(column_index), index)].alignment = alignment
                        ws['{}{}'.format(self.get_char(column_index), index)].border = border
                        ws['{}{}'.format(self.get_char(column_index + 1), index)] = columns[column_index+1]
                        ws['{}{}'.format(self.get_char(column_index + 1), index)].alignment = alignment
                        ws['{}{}'.format(self.get_char(column_index + 1), index)].border = border
                        ws['{}{}'.format(self.get_char(column_index + 2), index)] = columns[column_index+2]
                        ws['{}{}'.format(self.get_char(column_index + 2), index)].alignment = alignment
                        ws['{}{}'.format(self.get_char(column_index + 2), index)].border = border
                        ws['{}{}'.format(self.get_char(column_index + 3), index)] = columns[column_index+3]
                        ws['{}{}'.format(self.get_char(column_index + 3), index)].alignment = alignment
                        ws['{}{}'.format(self.get_char(column_index + 3), index)].border = border
                        ws['{}{}'.format(self.get_char(column_index + 4), index)] = columns[column_index+4]
                        ws['{}{}'.format(self.get_char(column_index + 4), index)].alignment = alignment
                        ws['{}{}'.format(self.get_char(column_index + 4), index)].border = border
                        index += 1

                        track_index = 1
                        if len(users_distances) == 4:
                            tracks = [index, index+3, index+2, index+1]
                        elif len(users_distances) == 3:
                            tracks = [index, index+1, index+2, index+3]
                        else:
                            tracks = [index, index+1, index+2, index+3]
                        for user_distance in users_distances:
                            ws['{}{}'.format(self.get_char(column_index - 1), tracks[track_index-1])] = user_distance.user.full_name
                            ws['{}{}'.format(self.get_char(column_index - 1), tracks[track_index-1])].border = border
                            ws['{}{}'.format(self.get_char(column_index), tracks[track_index-1])] = '{}({})'.format(user_distance.user.profile.get_age_group(), user_distance.user.profile.get_age_group_numbers())
                            ws['{}{}'.format(self.get_char(column_index), tracks[track_index-1])].border = border
                            try:
                                team = TeamRelationToUser.objects.filter(user=user_distance.user).first().team
                                CompetitionTeam.objects.get(team=team, competition=self.competition, is_complete=True)
                                team = team.name
                            except:
                                team = user_distance.user.profile.default_team
                            ws['{}{}'.format(self.get_char(column_index + 1), tracks[track_index-1])] = team
                            ws['{}{}'.format(self.get_char(column_index + 1), tracks[track_index-1])].border = border
                            ws['{}{}'.format(self.get_char(column_index + 2), tracks[track_index-1])] = user_distance.user.profile.city
                            ws['{}{}'.format(self.get_char(column_index + 2), tracks[track_index-1])].border = border
                            ws['{}{}'.format(self.get_char(column_index + 3), tracks[track_index-1])] = time_to_str(user_distance.pre_time)
                            ws['{}{}'.format(self.get_char(column_index + 3), tracks[track_index-1])].border = border
                            ws['{}{}'.format(self.get_char(column_index + 4), index + track_index -1)] = track_index
                            ws['{}{}'.format(self.get_char(column_index + 4), index + track_index -1)].border = border
                            track_index += 1
                        index += 5
                        swim_index += 1
                        distance_swim_index += 1
                distance_index += 1
        path = settings.BASE_DIR + "/media/predictions/" + str(self.competition.id) + ".xlsx"
        wb.save(path)
        return path

    def is_group_exists(self, group, members):
        for member in members:
            if member.profile.get_age_group() == group:
                return True

        return False

    def get_char(self, index):
        try:
            return self.alf[index]
        except:
            index -= 26
            return 'A' + self.alf[index]

    def get_all_members(self):
        competition_rel_user = CompetitionUser.objects.filter(competition=self.competition).all()
        single_members = list()
        for rel in competition_rel_user:
            single_members.append(rel.user)

        teams = CompetitionTeam.objects.filter(competition=self.competition).all()
        team_members = list()

        for team in teams:
            rels = TeamRelationToUser.objects.filter(team=team.team).all()
            for rel in rels:
                team_members.append(rel.user)

        return single_members + list(set(team_members) - set(single_members))
